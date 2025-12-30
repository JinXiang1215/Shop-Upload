import io
import pandas as pd
import re
import os
import requests
import json
import sys
import warnings
import shutil
import mimetypes
import xlsxwriter
import numpy as np
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
warnings.filterwarnings("ignore")

app = FastAPI()
# @app.post("/upload")
@app.post("/upload/{recordId}")
async def upload_file(recordId: str,file: UploadFile = File(...)):
    try:
        print("Upload triggered")
        print(f"Upload triggered for recordId: {recordId}")
        print(f"Uploaded file name is:{file.filename}")
        
        contents = await file.read()
        data = pd.read_excel(io.BytesIO(contents))

        # API responses
        api_call = f"https://rev.justtapao.com/jw/web/json/plugin/com.lineclear.GetShipmentType/service?id={recordId}"
        print(f"Making request to: {api_call}")

        response = requests.get(api_call)
        response_json = response.json()
        print("API response:", response_json)

        if isinstance(response_json, list) and len(response_json) > 0:
            first_item = response_json[0]
            platform_value=first_item.get("c_select_platform","")
            shipment_value=first_item.get("c_shipment_type","")
            oms_id=first_item.get("c_oms_accountNom","")

            print(f"Platform Value: {platform_value}")
            print(f"Shipment Value: {shipment_value}")
            print(f"OMS ID: {oms_id}")
        else:
            print("Error: API response is empty or not in expected format.")


        if not response_json:
            raise HTTPException(status_code=500, detail="Shipment Type API returned empty response. Please check API or ID.")

        # SKU to package ID mapping from API
        mapping_api = 'http://justtapao.com:8080/jw/web/json/plugin/com.lineclear.SKUPackageAPI/service?omsId=' + oms_id
        mapping_response = requests.get(mapping_api).json()
        if mapping_response:
            oms_id = mapping_response[0]['omsAccountNom']
            sku_to_packageid = {item['skuReferenceNo']: item['packageId'] for item in mapping_response}
        else:
            sku_to_packageid = {}


        # Template mapping from API 
        column_api = 'http://justtapao.com:8080/jw/web/json/plugin/com.lineclear.TemplateAPI/service?refId=' + platform_value
        column_response = requests.get(column_api).json()
        if not column_response:
            raise HTTPException(status_code=500, detail="Template API returned empty response. Please check platform refId.")
        print("response returned is:",column_response)

        def validate_dataframe(df, columns_to_check, reference_column):
            validation_results = []
            
            dropped_references = df[df.duplicated('Reference_Number', keep='last') & df['Reference_Number'].notna()]
            
            for index, row in df.iterrows():
                reference_number = row.get('Reference_Number', 'N/A')
                
                if index in dropped_references.index:
                    validation_results.append({
                        'Row': str(index + 2), 
                        'status': 'Dropped',
                        'description': f"Reference_Number {reference_number} is duplicated"
                    })
                else:
                    validation_result = validate_row(row, columns_to_check)
                    validation_results.append({
                        'Row': str(index + 2), 
                        'status': validation_result['status'], 
                        'description': validation_result['description']
                    })
            
            return json.dumps(validation_results, indent=4)


        def validate_row(row, columns_to_check):
            errors = [f'{column} is missing' for column in columns_to_check if pd.isna(row[column]) or row[column] == '']
            return {
                'status': 'Failed' if errors else 'Success',
                'description': '; '.join(errors) if errors else 'Data completed'
            }


        def combine_shipping_address(df, shipping_columns):
            df[shipping_columns[0]] = df[shipping_columns].apply(lambda row: ', '.join(row.dropna().astype(str)), axis=1)
            df.drop(columns=shipping_columns[1:], inplace=True)
            return df

        def filter_columns(column, address_column):
            address_columns_to_remove = address_column[1:]
            return [col for col in column if col not in address_columns_to_remove]

        def clean_data_symbols(df, column_names):
            symbols = ['#', '-', '(', ')', ' ', '+','.']
            for col in column_names:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str)
                    for symbol in symbols:
                        df[col] = df[col].str.replace(symbol, '', regex=False)
            return df

        column_name = ['Types_Of_Services', 'Service_Type', 'Packaging_Id', 'Delivery_Name', 'Delivery_Building_Unit','Unit_Number','Delivery_Address', 'Delivery_Address_2', 'Delivery_Postal_Code', 'Delivery_Phone_Number', 'Delivery Company Name', 'Shipment_Type', 'Mode', 'Reference_Number', 'Quantity', 'Weight', 'Length', 'Width', 'Height', 'DO_Numbers', 'ProtectionPlan_Product_Description', 'ProtectionPlan_Quantity', 'Total_Value', 'Parcel_Value', 'Sales_Tax', 'Delivery_Charges', 'COD', 'SST', 'Others']

        columns = [item for sublist in column_response.values() for item in (sublist if isinstance(sublist, list) else [sublist])]

        if not columns:
            raise HTTPException(status_code=500, detail="No columns received from template API. Cannot process file.")
        print(f"Columns from API: {columns}")
        print("DataFrame columns:", data.columns.tolist())
        data=data[columns]

        sku = column_response['Packaging_Id']
        data[sku] = data[sku].map(sku_to_packageid)
        data_column = data.columns
        ulive_column = {
            "Delivery_Address": "Customer",
            "Delivery_Name": "Customer",
            "Delivery_Phone_Number": "Customer",
            "Delivery_Postal_Code": "Customer",
            "Packaging_Id": "Product SKU",
            "Reference_Number": "Order Number"
        }

        if column_response == ulive_column:
            data.dropna(subset=['Order Number', 'Customer'], inplace=True)

        #if another template used, combines multiple adress fields into 1, clean numbers removing special characters
            def extract_info(cell_value):
                if not isinstance(cell_value, str):
                    cell_value = str(cell_value) if not pd.isna(cell_value) else ''
                
                pattern = r'^(.+?) \(\+(\d+)\) (.+)$'
                postcode_pattern = r'\b\d{5,6}\b'
                match = re.match(pattern, cell_value)
                if match:
                    name, phone, address = match.group(1), match.group(2), match.group(3)
                    postcode = re.findall(postcode_pattern, address)[0] if re.findall(postcode_pattern, address) else ''
                    return address, name, phone, postcode
                return cell_value, None, cell_value, None

            delivery_details = [key for key, value in column_response.items() if value == 'Customer']
            data[delivery_details] = data['Customer'].apply(extract_info).apply(pd.Series)
            filtered_column = [column_response['Reference_Number']] + delivery_details + ['Packaging_Id']

        else:
            address_column = column_response['Delivery_Address']
            if isinstance(address_column, list):
                data = combine_shipping_address(data, address_column)
            phone_column = column_response['Delivery_Phone_Number']
            data = clean_data_symbols(data, [phone_column])
            filtered_column = filter_columns(data_column, address_column)

        reverse_mapping = {}
        for key, value in column_response.items():
            if isinstance(value, list):
                for v in value:
                    reverse_mapping[v] = key
            else:
                reverse_mapping[value] = key

        rename_mapping = {col: reverse_mapping.get(col, col) for col in filtered_column} 
        data.rename(columns=rename_mapping, inplace=True)
        columns_with_defaults = {
        'Types_Of_Services': 'Domestic Shipment',
        'Delivery_Building_Unit': '-',
        'Delivery_Address_2': '-',
        'Unit_Number':'-',
        'Length': 1,
        'Width': 1,
        'Height': 1,
        'Quantity': 1,
        'Weight': 1,
        'Shipment_Type': 'Package',
        'Mode': shipment_value}

        for col, value in columns_with_defaults.items():
            data[col] = value

        def format_postal_code(code):
            if pd.isna(code) or code == '':
                return ''
            try:
                return str(int(float(code))).zfill(5)
            except ValueError:
                return ''

        data['Delivery_Postal_Code'] = data['Delivery_Postal_Code'].apply(format_postal_code)

        columns_to_check = ['Reference_Number', 'Delivery_Name', 'Delivery_Phone_Number', 'Delivery_Address', 'Delivery_Address_2', 'Delivery_Postal_Code']

        print(data['Delivery_Phone_Number'].head())
        validation_results = validate_dataframe(data, columns_to_check, sku)
        print(validation_results)
        
        data = data.drop_duplicates('Reference_Number', keep='last')
        for col in column_name:
            if col not in data.columns:
                data[col] = np.nan
        data = data[column_name]
        print(data)

        print("recordId:", recordId)
        
        if isinstance(validation_results, str):
            validation_results=json.loads(validation_results)

        print(f"validation_results type: {type(validation_results)}")  # Should print <class 'list'>
        print(f"First record type: {type(validation_results[0])}")  # Should print <class 'dict'>


        url= "http://justtapao.com:8080/jw/web/json/plugin/com.lineclear.getUpdateAPI/service"
        params = {'id':recordId}
        headers = {'Content-Type':'application/json'}

        r=requests.post(
            url=url,
            params=params,
            headers=headers,
            json=validation_results

        )
        
        print("Response from Joget:", r.status_code, r.text)

        print("Request body (debug):", json.dumps(validation_results, indent=2))

        data = data.fillna("#NUM!")

        # Save processed DataFrame to a new Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name="ProcessedData")

            workbook = writer.book
            worksheet = writer.sheets['ProcessedData']

            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

            header_font = Font(bold=True,size=12)
            cell_font = Font(size=10)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            header_fill = PatternFill(start_color='D9D9D9',end_color='D9D9D9',fill_type='solid')

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
                for cell in row:
                    cell.font = cell_font
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left')

        output.seek(0)

        headers = {
        'Content-Disposition': 'attachment; filename="processed_result.xlsx"'
    }
        
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers=headers)
      
    except Exception as e:
        print("ERROR:", e)
        return {"error": str(e)}
    